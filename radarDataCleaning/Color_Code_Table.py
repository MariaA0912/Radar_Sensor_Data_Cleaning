import pandas as pd
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Load CSV file into pandas DataFrame
df = pd.read_csv("cleaned_output.csv")

# Function to generate random pastel colors
def get_pastel_color():
    return f"{random.randint(200, 255):02X}{random.randint(200, 255):02X}{random.randint(200, 255):02X}"

# Generate unique colors for each unique value in Time, id, and detection_level
time_colors = {time: get_pastel_color() for time in df["Time"].unique()}
id_colors = {id_: get_pastel_color() for id_ in df["id"].unique()}
detection_colors = {dl: get_pastel_color() for dl in df["detection_level"].unique()}

# Create an Excel workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Radar Data"

# Write headers
ws.append(df.columns.tolist())

# Write data with color coding
for _, row in df.iterrows():
    # Assign colors based on unique values
    time_fill = PatternFill(start_color=time_colors[row["Time"]], end_color=time_colors[row["Time"]], fill_type="solid")
    id_fill = PatternFill(start_color=id_colors[row["id"]], end_color=id_colors[row["id"]], fill_type="solid")
    detection_fill = PatternFill(start_color=detection_colors[row["detection_level"]], end_color=detection_colors[row["detection_level"]], fill_type="solid")

    # Append row to Excel
    row_values = row.tolist()
    ws.append(row_values)

    # Apply color styling to each column
    for i, cell in enumerate(ws[ws.max_row], start=1):
        if i == 1:  # Time column
            cell.fill = time_fill
        elif i == 2:  # ID column
            cell.fill = id_fill
        elif i == 3:  # Detection Level column
            cell.fill = detection_fill

# Save the Excel file
wb.save("color_coded_table.xlsx")

print("Excel file with color-coded table created successfully!")
