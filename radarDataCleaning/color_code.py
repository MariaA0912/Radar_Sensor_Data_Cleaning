import csv
import openpyxl
from openpyxl.styles import PatternFill

# Function to find cells with duplicates and similar numbers
def highlight_duplicates_and_similar_numbers(input_file, output_file):
    # Open CSV file and load data
    with open(input_file, mode='r', newline='', encoding='utf-8') as infile:
        reader = csv.reader(infile)
        data = list(reader)
    
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write the CSV data to the Excel worksheet
    for row_num, row in enumerate(data, start=1):
        for col_num, cell in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_num, value=cell)

    # Color fill for highlighted cells
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Step 1: Find and color duplicate words
    word_counts = {}
    for row in data:
        for cell in row:
            if cell.isalpha():  # Check if it's a word
                word_counts[cell] = word_counts.get(cell, 0) + 1
    
    # Apply color to duplicates in words
    for row_num, row in enumerate(data, start=1):
        for col_num, cell in enumerate(row, start=1):
            if cell.isalpha() and word_counts.get(cell, 0) > 1:
                ws.cell(row=row_num, column=col_num).fill = highlight_fill

    # Step 2: Find and color cells with similar numbers
    def are_numbers_similar(num1, num2, tolerance=0.01):
        # Compare two numbers up to 2 decimal places (within a given tolerance)
        return abs(float(num1) - float(num2)) <= tolerance

    for row_num in range(len(data)):
        for col_num in range(len(data[row_num])):
            for compare_row_num in range(len(data)):
                for compare_col_num in range(len(data[compare_row_num])):
                    if row_num != compare_row_num or col_num != compare_col_num:
                        try:
                            if are_numbers_similar(data[row_num][col_num], data[compare_row_num][compare_col_num]):
                                # Apply color to similar numbers
                                ws.cell(row=row_num+1, column=col_num+1).fill = highlight_fill
                                ws.cell(row=compare_row_num+1, column=compare_col_num+1).fill = highlight_fill
                        except ValueError:
                            pass  # Skip cells that aren't numbers

    # Save the modified file to the output file
    wb.save(output_file)

# Example usage
highlight_duplicates_and_similar_numbers('clean_data.csv', 'highlighted_output.xlsx')
